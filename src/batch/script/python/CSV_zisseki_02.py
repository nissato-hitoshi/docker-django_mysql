class ParamError(Exception):
    pass

try:
    import openpyxl
    import sys
    import pandas as pd
    import datetime
    import csv
    import logging
    import os
    import MySQLdb
    import config       # config.py モジュール（自作）


    ######################################
    # 関数                               #
    ######################################
    
    # 実行ファイル名
    f = os.path.basename(__file__)
    file_name = os.path.splitext(f)[0]  # 拡張子なしファイル名

    def getMyLogger(name, file_name):
        """
        ログ出力

        Parameters
        ----------
        name : str
        file_name : str

        Returns
        -------
        logger

        """
        logging.basicConfig(level=logging.DEBUG)
        logger = logging.getLogger(name)
        logger.setLevel(logging.DEBUG)
        handler = logging.FileHandler(config.LOG_PATH + file_name + '.log')
        handler.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(levelname)-9s  %(asctime)s  [%(name)s] %(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        return logger
 
    def getsuzi_func(ws, row_num, kaikeiki):
        """
        月次実績処理

        Parameters
        ----------
        ws : Worksheet
        row_num : int
        kaikeiki : int
        """
        # プロジェクトコード
        if ws.cell(row = row_num, column = 1).value is None:
            pj_code = ''
        else:
            pj_code_v = ws.cell(row = row_num, column = 1).value.replace('\n','')
            pj_code = pj_code_v.strip()
        # プロジェクト名
        if ws.cell(row = row_num, column = 2).value is None:
            pj_name = ''
        else:
            pj_name = ws.cell(row = row_num, column = 2).value.replace('\n','')
        # 取引先
        if ws.cell(row = row_num, column = 3).value is None:
            torihiki = ''
        else:
            torihiki = ws.cell(row = row_num, column = 3).value.replace('\n','')
        # 計上月
        if ws.cell(row = row_num, column = 4).value is None:
            keizyou = ''
        else:
            value = ws.cell(row = row_num, column = 4).value
            m = value[:-1]
            if len(m) == 1:                 # 1月～9月の場合
                m = '0' + m                 # mmの形にする
                y = str(kaikeiki + 1979)
            else:                           # 10月～12月の場合
                y = str(kaikeiki + 1978)
            keizyou = y + m
        # 営業部門
        if ws.cell(row = row_num, column = 5).value is None:
            eigyou = ''
        else:
            eigyou = ws.cell(row = row_num, column = 5).value.replace('\n','')
        # 事業部
        if ws.cell(row = row_num, column = 6).value is None:
            zigyou = ''
        else:
            zigyou = ws.cell(row = row_num, column = 6).value.replace('\n','')
        # 主管部門
        if ws.cell(row = row_num, column = 7).value is None:
            syukan = ''
        else:
            syukan = ws.cell(row = row_num, column = 7).value.replace('\n','')
        # 売上高
        if ws.cell(row = row_num, column = 8).value is None:
            uriage = ''
        else:
            uriage = ws.cell(row = row_num, column = 8).value
        # 材料費
        if ws.cell(row = row_num, column = 9).value is None:
            zairyou = ''
        else:
            zairyou = ws.cell(row = row_num, column = 9).value
        # 労務費
        if ws.cell(row = row_num, column = 10).value is None:
            gaityu = ''
        else:
            gaityu = ws.cell(row = row_num, column = 10).value
        # 外注費
        if ws.cell(row = row_num, column = 11).value is None:
            roumu = ''
        else:
            roumu = ws.cell(row = row_num, column = 11).value
        # 経費
        if ws.cell(row = row_num, column = 12).value is None:
            keihi = ''
        else:
            keihi = ws.cell(row = row_num, column = 12).value
        # 原価計
        if ws.cell(row = row_num, column = 13).value is None:
            genka = ''
        else:
            genka = ws.cell(row = row_num, column = 13).value
        # 粗利額
        if ws.cell(row = row_num, column = 14).value is None:
            sorigaku = ''
        else:
            sorigaku = ws.cell(row = row_num, column = 14).value
        # 粗利率
        if ws.cell(row = row_num, column = 15).value is None:
            soriritsu = ''
        else:
            soriritsu = ws.cell(row = row_num, column = 15).value
        # 契約形態
        if ws.cell(row = row_num, column = 16).value is None:
            keiyaku = ''
        else:
            keiyaku = ws.cell(row = row_num, column = 16).value.replace('\n','')
        # 種別
        if ws.cell(row = row_num, column = 17).value is None:
            syubetsu = ''
        else:
            syubetsu = ws.cell(row = row_num, column = 17).value.replace('\n','')
        # プロジェクト親番
        if ws.cell(row = row_num, column = 18).value is None:
            pj_p_code = ''
        else:
            pj_p_code_value = ws.cell(row = row_num, column = 18).value
            pj_p_code = str(pj_p_code_value)
        #プロジェクト親番名
        if ws.cell(row = row_num, column = 19).value is None:
            pj_p_name = ''
        else:
            pj_p_name = ws.cell(row = row_num, column = 19).value.replace('\n','')

        
        # リストに追加
        accounting_period.append(kaikeiki)          # 会計期
        project_code.append(pj_code)               # プロジェクトコード
        project_name.append(pj_name)               # プロジェクト名
        client_name.append(torihiki)                # 取引先
        accounting_month.append(keizyou)           # 計上月
        sales_department.append(eigyou)           # 営業部門
        business_sector.append(zigyou)            # 事業部
        supervising_department.append(syukan)     # 主管部門
        mount_of_sales.append(uriage)             # 売上高
        material_cost.append(zairyou)              # 材料費
        labor_cost.append(roumu)                 # 労務費
        outsourcing_cost.append(gaityu)           # 外注費
        expenses.append(keihi)                   # 経費
        cost_total.append(genka)                 # 原価計
        gross_profit.append(sorigaku)               # 粗利額
        gross_profit_margin.append(soriritsu)        # 粗利率
        type_of_contract.append(keiyaku)           # 契約形態
        kinds.append(syubetsu)                      # 種別
        project_parent_code.append(pj_p_code)        # プロジェクト親番
        project_parent_name.append(pj_p_name)        # プロジェクト親番名

    def kadou(ws, row_num, kaikeiki, file_ym):
        """
        稼働準備金処理

        Parameters
        ----------
        ws : Worksheet
        row_num : int
        kaikeiki : int
        file_ym : int
        """
        # プロジェクトコード
        if ws.cell(row = row_num, column = 1).value is None:
            pj_code = ''
        else:
            pj_code_v = ws.cell(row = row_num, column = 1).value.replace('\n','')
            pj_code = pj_code_v.strip()
        # プロジェクト名
        if ws.cell(row = row_num, column = 2).value is None:
            pj_name = ''
        else:
            pj_name = ws.cell(row = row_num, column = 2).value.replace('\n','')
        # 取引先
        torihiki = '-'
        # 計上月
        keizyou = file_ym
        # 営業部門
        eigyou = '-'
        # 事業部
        a = 0
        A9 = ['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14']
        z_arr = ['ZZ', 'SD', 'IS', 'LCM', 'PA', 'SD1', 'SD2', 'SD3', 'SD4', 'IS1', 'IS2', 'IS3', 'LCM', 'PA1', 'PA2']
        for a in range(15):
            if pj_code[-2:] == A9[a]:
                zigyou = z_arr[a]
                break
            else:
                zigyou = 'XX'
        # 主管部門
        a = 0
        A9 = ['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14']
        s_arr = ['ZZ', 'SD', 'IS', 'LCM', 'PA', 'SD', 'SD', 'SD', 'SD', 'IS', 'IS', 'IS', 'LCM', 'PA', 'PA']
        for a in range(15):
            if pj_code[-2:] == A9[a]:
                syukan = s_arr[a]
                break
            else:
                syukan = 'XX'
        # 売上高
        uriage = 0
        # 材料費
        zairyou = 0
        # 労務費
        if ws.cell(row = row_num, column = 3).value is None:
            roumu = ''
        else:
            roumu = ws.cell(row = row_num, column = 3).value
        # 外注費
        gaityu = 0
        # 経費
        keihi = 0
        # 原価計
        if ws.cell(row = row_num, column = 3).value is None:
            genka = ''
        else:
            genka = ws.cell(row = row_num, column = 3).value
        # 粗利額
        if ws.cell(row = row_num, column = 3).value is None:
            sorigaku = ''
        else:
            sorigaku_v = ws.cell(row = row_num, column = 3).value
            sorigaku = int(sorigaku_v) * 1
        # 粗利率
        if  ws.cell(row = row_num, column = 3).value is None:
            soriritsu = 0
        elif ws.cell(row = row_num, column = 3).value >= 0:
            soriritsu = -100
        else:
            soriritsu = 0
        # 契約形態
        keiyaku = '-'
        # 種別
        syubetsu = '-'
        # プロジェクト親番
        pj_p_code = '38300015'
        # プロジェクト親番名
        pj_p_name = '技術研修・稼動準備'

        # リストに追加
        accounting_period.append(kaikeiki)          # 会計期
        project_code.append(pj_code)               # プロジェクトコード
        project_name.append(pj_name)               # プロジェクト名
        client_name.append(torihiki)                # 取引先
        accounting_month.append(keizyou)           # 計上月
        sales_department.append(eigyou)           # 営業部門
        business_sector.append(zigyou)            # 事業部
        supervising_department.append(syukan)     # 主管部門
        mount_of_sales.append(uriage)             # 売上高
        material_cost.append(zairyou)              # 材料費
        labor_cost.append(roumu)                 # 労務費
        outsourcing_cost.append(gaityu)           # 外注費
        expenses.append(keihi)                   # 経費
        cost_total.append(genka)                 # 原価計
        gross_profit.append(sorigaku)               # 粗利額
        gross_profit_margin.append(soriritsu)        # 粗利率
        type_of_contract.append(keiyaku)           # 契約形態
        kinds.append(syubetsu)                      # 種別
        project_parent_code.append(pj_p_code)        # プロジェクト親番
        project_parent_name.append(pj_p_name)        # プロジェクト親番名

    def mysql_connect():
        """
        MySQL接続

        Returns
        -------
        conn : Connection
        cur : Cursor
        """
        conn = MySQLdb.connect(
            user=config.MYSQL_USER,
            port=int(config.MYSQL_PORT),
            passwd=config.MYSQL_PW,
            host=config.MYSQL_HOST,
            db=config.MYSQL_DB,
            local_infile=True)

        cur = conn.cursor() # カーソルを取得する
        return (conn, cur)

    def csv_load(kaikeiki, output_name):
        """
        CSVへのLOAD関数

        Parameters
        ----------
        kaikeiki : int
        output_name : str

        """
        conn, cur = mysql_connect() # MySQL接続
        try:
            # 対象期のデータを削除
            delete_sql = 'DELETE FROM sales_achievement WHERE accounting_period = %s;'
            cur.execute(delete_sql, (kaikeiki,))
            # 列指定LOAD
            load_sql = ('''LOAD DATA LOCAL INFILE %s
                INTO TABLE sales_achievement
                FIELDS TERMINATED BY ','
                ENCLOSED BY '"'
                LINES TERMINATED BY '\r\n'
                IGNORE 1 LINES
                (@1, @2, @3, @4, @5, @6, @7, @8, @9, @10, @11, @12, @13, @14, @15, @16, @17, @18, @19, @20)
                SET
                updated = now(),
                created = now(),
                accounting_period = @1,
                project_code = @2,
                project_name = @3,
                client_name = @4,
                accounting_month = @5,
                sales_department = @6,
                business_sector = @7,
                supervising_department = @8,
                mount_of_sales = @9,
                material_cost = @10,
                labor_cost = @11,
                outsourcing_cost = @12,
                expenses = @13,
                cost_total = @14,
                gross_profit = @15,
                gross_profit_margin = @16,
                type_of_contract = @17,
                kinds = @18,
                project_parent_code = @19,
                project_parent_name = @20''')
            cur.execute(load_sql, (output_name,))
            conn.commit()
        except MySQLdb.Error as e:
            print('MySQLdb.Error: ', e)
        cur.close
        conn.close



    ######################################
    # メイン処理                         #
    ######################################

    # パラメータ
    args = sys.argv

    # 会計期処理
    if args[2] is None:     # パラメータの指定がない場合はエラー
        raise ParamError
    else:
        year_month = args[2]
    year = year_month[0:4]  # 年 yyyy
    month = year_month[4:6] # 月 mm

    # 10、11、12月の場合
    if month == '10' or month == '11' or month == '12':
        kaikeiki = int(year) - 1978
        input_month = month
    # それ以外の月の場合
    elif month == '01' or month == '02' or month == '03' or month == '04' or month == '05' or month == '06' or month == '07' or month == '08' or month == '09':
        kaikeiki = int(year) - 1979
        input_month = month[1]


    # リスト初期化
    accounting_period = []          # 会計期
    project_code = []               # プロジェクトコード
    project_name = []               # プロジェクト名
    client_name = []                # 取引先
    accounting_month = []           # 計上月
    sales_department = []           # 営業部門
    business_sector = []            # 事業部
    supervising_department = []     # 主管部門
    mount_of_sales = []             # 売上高
    material_cost = []              # 材料費
    labor_cost = []                 # 労務費
    outsourcing_cost = []           # 外注費
    expenses = []                   # 経費
    cost_total = []                 # 原価計
    gross_profit = []               # 粗利額
    gross_profit_margin = []        # 粗利率
    type_of_contract = []           # 契約形態
    kinds = []                      # 種別
    project_parent_code = []        # プロジェクト親番
    project_parent_name = []        # プロジェクト親番名

    # 月次実績読み込み 
    input_file_1 = input_month + '月月次実績データ.xlsx'
    TargetPath1 = config.ZISSEKI_PATH + args[1] + '期' + '/' + args[2] + '/' + input_file_1
    wb1 = openpyxl.load_workbook(TargetPath1)
    sheet_names1 = wb1.sheetnames

    # 月次実績配列ループ
    for i in range(0,len(sheet_names1)):
        ws_name1 = sheet_names1[i]
        if ws_name1 != '月次実績データ':
            continue
        ws1 = wb1[ws_name1]

        # Excelファイル2行目以降を1行ずつ読み込んで、項目ごとのリストに追加
        for row in ws1.iter_rows(min_row=2):
            for cell in row:
                row_num = cell.row  # 行番号
                a = ws1.cell(row = row_num, column = 1).value    # A列
                if a == '':
                    continue
                getsuzi_func(ws1, row_num, kaikeiki)

    # 稼動準備金読み込み
    m_arr = ['10', '11', '12', '01', '02', '03', '04', '05', '06', '07', '08', '09']
    idx = m_arr.index(month)
    for e in range(idx + 1):
        # 10、11、12月の場合
        if idx == 0 or idx == 1 or idx == 2:
            y = str(kaikeiki + 1978)
            file_ym = y + m_arr[e]
            input_file_2 = '主管部門別稼動準備金額' + m_arr[e] + '月.xlsx'
        # それ以外の月の場合
        else:
            if e == 0 or e == 1 or e == 2:
                y = str(kaikeiki + 1978)
                file_ym = y + m_arr[e]
                input_file_2 = '主管部門別稼動準備金額' + m_arr[e] + '月.xlsx'
            else:
                y = str(kaikeiki + 1979)
                file_ym = y + m_arr[e]
                input_file_2 = '主管部門別稼動準備金額' + m_arr[e][1] + '月.xlsx'

        TargetPath2 = config.ZISSEKI_PATH + args[1] + '期' + '/' + file_ym + '/' + input_file_2  # 読み込むファイルのパス
        if not os.path.exists(TargetPath2):     # パスが存在しない場合
            continue
        wb2 = openpyxl.load_workbook(TargetPath2)
        sheet_names2 = wb2.sheetnames

        # 稼動準備金 配列ループ
        for i in range(0,len(sheet_names2)):
            ws_name2 = sheet_names2[i]
            if ws_name2 != 'プロジェクト一覧表':
                continue
            ws2 = wb2[ws_name2]

            # Excelファイル9行目以降を1行ずつ読み込んで、項目ごとのリストに追加
            for row in ws2.iter_rows(min_row=9):
                for cell in row:
                    row_num = cell.row  # 行番号
                    a = ws2.cell(row = row_num, column = 1).value    # A列
                    # A列が空白（最終行）のとき
                    if a == '':
                        continue
                    kadou(ws2, row_num, kaikeiki, file_ym)


    # 出力用辞書型
    dict1 = { 'accounting_period': accounting_period,
            'project_code': project_code, 
            'project_name': project_name,
            'client_name': client_name,
            'accounting_month': accounting_month,
            'sales_department': sales_department,
            'business_sector': business_sector,
            'supervising_department': supervising_department,
            'mount_of_sales': mount_of_sales,
            'material_cost': material_cost,
            'labor_cost': labor_cost,
            'outsourcing_cost': outsourcing_cost,
            'expenses': expenses,
            'cost_total': cost_total,
            'gross_profit': gross_profit,
            'gross_profit_margin': gross_profit_margin,
            'type_of_contract': type_of_contract,
            'kinds': kinds,
            'project_parent_code': project_parent_code,
            'project_parent_name': project_parent_name
            }

    #dataframe作成
    df = pd.DataFrame(dict1)

    # 出力ファイル用日付
    t_delta = datetime.timedelta(hours=9)
    JST = datetime.timezone(t_delta, 'JST')
    now = datetime.datetime.now(JST)
    d = now.strftime('%Y%m')


    # 文字コード、区切り文字処理
    if len(args) >= 6:
        # 文字コード
        if args[4] == ' ':
            my_encoding = 'utf_8_sig'
        else:
            my_encoding = args[4]
        # 区切り文字
        my_sep = args[5]
    elif len(args) == 5:
        # 文字コード
        if args[4] == ' ':
            my_encoding = 'utf_8_sig'
        else:
            my_encoding = args[4]
        # 区切り文字
        my_sep = ','
    else:
        # 文字コード
        my_encoding = 'utf_8_sig'
        # 区切り文字
        my_sep = ','

    # 出力ファイル名
    output_name = config.OUTPUT_PATH + str(kaikeiki) + '_' + d + '_sales_achievement.csv'

    # CSV形式で出力
    df.to_csv(output_name, sep=my_sep, encoding=my_encoding, index=False, quoting=csv.QUOTE_NONNUMERIC)

    # CSVからLOAD
    if int(args[3]) == 1:    # LOADするとき
        csv_load(kaikeiki, output_name)
    elif int(args[3]) == 2:  # LOADしないとき
        pass
    else:
        raise ParamError

except(ImportError, ModuleNotFoundError) as e:
    print('Import Error!:', e)

except(OSError, SyntaxError, TypeError, UnicodeError) as e:
    print('Error!:', e)

except FileNotFoundError as e:
    print('FileNotFound!:', e)

except ParamError as e:
    print('Parameter Error!', e)

else:
    print('Succesfully End!')

finally:
    logger = getMyLogger(__name__, file_name)
    logger.debug('デバッグ')
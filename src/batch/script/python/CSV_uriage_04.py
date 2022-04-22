class ParamError(Exception):
    pass

try:
    import openpyxl
    import sys
    import pandas as pd
    import datetime
    import csv
    import logging
    import MySQLdb
    import os
    import config       # config.py モジュール（自作）

    ######################################
    # 関数                               #
    ######################################
    
    # 実行ファイル名
    f = os.path.basename(__file__)
    file_name = os.path.splitext(f)[0]  # 拡張子なしファイル名

    def getMyLogger(name, file_name):
        """
        ログ出力

        Parameters
        ----------
        name : str
        file_name : str

        Returns
        -------
        logger

        """
        logging.basicConfig(level=logging.DEBUG)
        logger = logging.getLogger(name)
        logger.setLevel(logging.DEBUG)
        handler = logging.FileHandler(config.LOG_PATH + file_name + '.log')
        handler.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(levelname)-9s  %(asctime)s  [%(name)s] %(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        return logger
 
    def sales_dept(ws_name):
        """
        分類処理

        Parameters
        ----------
        ws_name : str

        Returns
        -------
        bunrui : str
        """
        if '外注' in ws_name:
            bunrui = '2:外注'
        elif '物販' in ws_name:
            bunrui = '3:物販'
        else:
            bunrui = '1:社員'
        return bunrui
 
    def list_func(ws, row_num):
        """
        リスト処理

        Parameters
        ----------
        ws : Worksheet
        row_num : int
        """
        # E列 営業担当
        eigyou = ws.cell(row = row_num, column = 5).value
        if eigyou is None:
            eigyou = ''
        else:
            eigyou = eigyou.replace('\n','')

        # H列 主管部門
        syukann = ws.cell(row = row_num, column = 8).value
        if syukann is None:
            syukann = ''
        else:
            syukann = syukann.replace('\n','')

        # C列 顧客
        kokyaku = ws.cell(row = row_num, column = 3).value
        if kokyaku is None:
            kokyaku = ''
        else:
            kokyaku = kokyaku.replace('\n','')
        
        # F列 プロジェクト名
        pj = ws.cell(row = row_num, column = 6).value
        if pj is None:
            pj = ''
        else:
            pj = pj.replace('\n','')

        # I列 作業員
        sagyou = ws.cell(row = row_num, column = 9).value
        if sagyou is None:
            sagyou = ''
        else:
            sagyou = sagyou.replace('\n','')

        # リストに追加
        sales_representative.append(eigyou)
        supervising_department.append(syukann)
        client_name.append(kokyaku)
        project_name.append(pj)
        operator.append(sagyou)

    def mysql_connect():
        """
        MySQL接続

        Returns
        -------
        conn : Connection
        cur : Cursor
        """
        conn = MySQLdb.connect(
            user=config.MYSQL_USER,
            port=int(config.MYSQL_PORT),
            passwd=config.MYSQL_PW,
            host=config.MYSQL_HOST,
            db=config.MYSQL_DB,
            local_infile=True)

        cur = conn.cursor() # カーソルを取得する
        return (conn, cur)

    def version2(kaikeiki):
        """
        バージョン管理 関数2

        Parameters
        ----------
        kaikeiki : int

        Returns
        -------
        current_version : int
        
        """
        conn, cur = mysql_connect() # MySQL接続
        current_version = ""
        try:
            sql = "select count(current_version) from version_control where accounting_period = %s;"
            cur.execute(sql, (kaikeiki,))

            # 実行結果を取得する
            for row in cur.fetchone():            
                current_ver_count = row

            if current_ver_count == 0:
                ver2_insert = "insert into version_control values (0, %s, 1, 1, 1, cast( now() as datetime ), cast( now() as datetime ));"
                cur.execute(ver2_insert, (kaikeiki,))
                ver2_select = "select current_version from version_control where accounting_period = %s;"
                cur.execute(ver2_select, (kaikeiki,))
                for row in cur.fetchone():            
                    current_version = row
                conn.commit()
            elif current_ver_count == 1:
                ver2_select = "select current_version from version_control where accounting_period = %s;"
                cur.execute(ver2_select, (kaikeiki,))
                for row in cur.fetchone():            
                    current_version = row
                current_update = current_version + 1
                ver2_update = "SET SQL_SAFE_UPDATES = 0; update version_control set current_version = %s, updated = cast( now() as datetime ) where accounting_period =%s;"
                cur.execute(ver2_update, (current_update, kaikeiki))
                cur.execute(ver2_select, (kaikeiki,))
                for row in cur.fetchone():            
                    current_version = row
                conn.commit()

        except MySQLdb.Error as e:
            print('MySQLdb.Error: ', e)

        cur.close
        conn.close
        return current_version

    def version3(kaikeiki):
        """
        バージョン管理 関数3

        Parameters
        ----------
        kaikeiki : int

        """
        conn, cur = mysql_connect() # MySQL接続
        try:
            ver3 = "select second_half_version from version_control where accounting_period = %s;"
            cur.execute(ver3, (kaikeiki,))
            for row in cur.fetchone():            
                result = row

            if result is None:
                print('error!')
            else:
                result = result + 1
                ver3_update = "update version_control set second_half_version = %s, updated = cast( now() as datetime ) where accounting_period = %s;"
                cur.execute(ver3_update, (result, kaikeiki))
                conn.commit()
        except MySQLdb.Error as e:
            print('MySQLdb.Error: ', e)
        cur.close
        conn.close

    def csv_load(output_name):
        """
        CSVへのLOAD関数

        Parameters
        ----------
        output_name : str

        """
        conn, cur = mysql_connect() # MySQL接続
        try:
            sql = ('''LOAD DATA LOCAL INFILE %s
                INTO TABLE sales_budget
                FIELDS TERMINATED BY ','
                ENCLOSED BY '"'
                LINES TERMINATED BY '\r\n'
                IGNORE 1 LINES
                (@1, @2, @3, @4, @5, @6, @7, @8, @9, @10, @11, @12, @13, @14, @15, @16)
                SET
                updated = now(),
                created = now(),
                accounting_period = @1,
                version = @2,
                sales_department = @3,
                classification = @4,
                sales_representative = @5,
                supervising_department = @6,
                client_name = @7,
                project_name = @8,
                operator = @9,
                accounting_month = @10,
                sales_budget = @11,
                sales_expected = @12,
                sales_achievement = @13,
                half_period = @14,
                quarter = @15,
                accuracy = @16''')
            cur.execute(sql, (output_name,))
            conn.commit()
        except MySQLdb.Error as e:
            print('MySQLdb.Error: ', e)
        cur.close
        conn.close


    ######################################
    # メイン処理                          #
    ######################################

    # 読み込み 
    args = sys.argv
    TargetPath = config.YOSAN_PATH + args[1] + '期売上予算_営業資料.xlsm'   # 読み込むExcelファイル名をパラメータから取得
    wb = openpyxl.load_workbook(TargetPath)
    sheet_names = wb.sheetnames

    # リスト初期化
    accounting_period = []      # 会計期
    version = []                # バージョン
    sales_department = []       # 営業部門
    classification = []         # 分類
    sales_representative = []   # 営業担当
    supervising_department = [] # 主管部門
    client_name = []            # 顧客
    project_name = []           # プロジェクト名
    operator = []               # 作業員
    accounting_month = []       # 年月
    sales_budget = []           # 予算
    sales_expected = []         # 見込み
    sales_achievement = []      # 実績
    half_period = []            # 半期
    quarter = []                # 四半期
    accuracy = []               # 確度

    # 会計期
    kaikeiki = int(args[1])  # パラメータから会計期を取得


    # バージョン
    if int(args[2]) == 1:    # 取込ファイルが期首予算、見込取込の場合
        ver = version2(kaikeiki)
    elif int(args[2]) == 2:  # 取込ファイルが下期予算確定バージョンの場合
        version3(kaikeiki)
        ver = version2(kaikeiki)
    else:
        raise ParamError


    # 配列ループ
    for i in range(0,len(sheet_names)):
        ws_name = sheet_names[i]
        ws = wb[ws_name]


        # 営業部門
        if '実績第一' in ws_name:
            eigyou_bumon = '1:第一'
        elif '実績第二' in ws_name:
            eigyou_bumon = '2:第二'
        elif '実績ＢＳ' in ws_name:
            eigyou_bumon = '3:BS'
        elif '実績ＳＩ' in ws_name:
            eigyou_bumon = '4:SI'   
        else:
            continue
        
        # 分類
        bunrui = sales_dept(ws_name)

        #Excelファイル5行目以降を1行ずつ読み込んで、項目ごとのリストに追加
        for row in ws.iter_rows(min_row=5):
            for cell in row:

                # 年月処理
                year1 = str(kaikeiki + 1978) #10月～12月
                year2 = str(kaikeiki + 1979) #1月～9月
                month_count = [10, 12, 14, 16, 18, 20, 24, 26, 28, 30, 32, 34]
                month_arr = [year1+'10', year1+'11', year1+'12', year2+'01', year2+'02', year2+'03', year2+'04', year2+'05', year2+'06', year2+'07', year2+'08', year2+'09']
                hanki_arr = ['1:上期', '1:上期', '1:上期', '1:上期', '1:上期', '1:上期', '2:下期', '2:下期', '2:下期', '2:下期', '2:下期', '2:下期',]
                shihanki_arr = ['1Q', '1Q', '1Q', '2Q', '2Q', '2Q', '3Q', '3Q', '3Q', '4Q', '4Q', '4Q']
                kakudo_count = [41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52]

                for a in range(12):
                    count = month_count[a]

                    row_num = cell.row  # 行番号
                    c = ws.cell(row = row_num, column = 3).value     # C列 顧客先の値
                    j = ws.cell(row = row_num, column = 10).value    # J列 売上見込の値
                        
                    # C列が空白、J列が空白でないとき（＝最終行のとき）
                    if (c is None) and (j is not None):
                        continue

                    if cell.col_idx == count:
                        col_num = cell.col_idx  # 列数
                        mikomi_num = ws.cell(row = row_num, column = col_num + 1).value #見込の値
                        
                        # 予算が空白、見込が空白のとき
                        if (cell.value is None) and (mikomi_num is None):
                            continue
                        # 予算に値があり、見込が空白のとき
                        elif (cell.value is not None) and (mikomi_num is None):
                            yosan = cell.value
                            mikomi = 0
                            zisseki = yosan
                        # 予算に値があり、見込に値があるとき
                        elif (cell.value is not None) and (mikomi_num is not None):
                            yosan = cell.value
                            mikomi = mikomi_num
                            zisseki = mikomi
                        else:
                            continue

                        ym = int(month_arr[a])          # 年月
                        hanki = hanki_arr[a]            # 半期
                        shihanki = shihanki_arr[a]      # 四半期
                        kakudo = ws.cell(row = row_num, column = kakudo_count[a]).value # 確度
                        if kakudo is None:
                            kakudo = ''

                        # リストに追加
                        accounting_period.append(kaikeiki)      # 会計期
                        version.append(ver)                     # バージョン
                        sales_department.append(eigyou_bumon)   # 営業部門
                        classification.append(bunrui)           # 分類
                        accounting_month.append(ym)       # 年月
                        sales_budget.append(yosan)        # 予算
                        sales_expected.append(mikomi)     # 見込
                        sales_achievement.append(zisseki) # 実績
                        half_period.append(hanki)   # 半期
                        quarter.append(shihanki)    # 四半期
                        accuracy.append(kakudo)     # 確度

                        list_func(ws, row_num)


    # 出力用辞書型
    dict1 = { 'accounting_period': accounting_period,
            'version': version, 
            'sales_department': sales_department,
            'classification': classification,
            'sales_representative': sales_representative,
            'supervising_department': supervising_department,
            'client_name': client_name,
            'project_name': project_name,
            'operator': operator,
            'accounting_month': accounting_month,
            'sales_budget': sales_budget,
            'sales_expected': sales_expected,
            'sales_achievement': sales_achievement,
            'half_period': half_period,
            'quarter': quarter,
            'accuracy': accuracy
            }

    #dataframe作成
    df = pd.DataFrame(dict1)

    # 出力ファイル用日付
    t_delta = datetime.timedelta(hours=9)
    JST = datetime.timezone(t_delta, 'JST')
    now = datetime.datetime.now(JST)
    d = now.strftime('%Y%m') # yyyymm式

    # パラメータ
    if len(args) >= 6:
        # 文字コード
        if args[4] == ' ':
            my_encoding = 'utf_8_sig'
        else:
            my_encoding = args[4]
        # 区切り文字
        my_sep = args[5]
    elif len(args) == 5:
        # 文字コード
        if args[4] == ' ':
            my_encoding = 'utf_8_sig'
        else:
            my_encoding = args[4]
        # 区切り文字
        my_sep = ','
    else:
        # 文字コード
        my_encoding = 'utf_8_sig'
        # 区切り文字
        my_sep = ','

    # 出力ファイル名
    output_name = config.OUTPUT_PATH + str(kaikeiki) + '_' + str(ver) + '_' + d + '_sales_budget.csv'

    # CSV形式で出力
    df.to_csv(output_name, sep=my_sep, encoding=my_encoding, index=False, quoting=csv.QUOTE_NONNUMERIC)

    # CSVからLOAD
    if int(args[3]) == 1:    # LOADするとき
        csv_load(output_name)
    elif int(args[3]) == 2:  # LOADしないとき
        pass
    else:
        raise ParamError

except(ImportError, ModuleNotFoundError) as e:
    print('Import Error!:', e)

except(OSError, SyntaxError, TypeError, UnicodeError) as e:
    print('Error!:', e)

except FileNotFoundError as e:
    print('FileNotFound!:', e)

except ParamError as e:
    print('Parameter Error!', e)

else:
    print('Succesfully End!')

finally:
    logger = getMyLogger(__name__, file_name)
    logger.debug('デバッグ')